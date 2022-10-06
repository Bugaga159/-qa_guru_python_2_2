import shutil
from zipfile import ZipFile
import pytest
from PyPDF2 import PdfReader
import os.path
import csv
from openpyxl import load_workbook

@pytest.fixture()
def create_zip():
    if os.path.exists('./resources/test.zip'):
        os.remove('./resources/test.zip')

    zip_archive = ZipFile('test.zip', 'w')
    zip_archive.write('./resources/pdf-test.pdf')
    zip_archive.write('./resources/SampleCSVFile_11kb.csv')
    zip_archive.write('./resources/file_example_XLSX_10.xlsx')
    zip_archive.close()
    shutil.move('test.zip', './resources')
    shutil.unpack_archive('./resources/test.zip', 'hidden')

def test_check_pdf(create_zip):
    pdf_reader = PdfReader('./hidden/resources/pdf-test.pdf')
    pdf_size = os.path.getsize('./hidden/resources/pdf-test.pdf')
    number_of_page = len(pdf_reader.pages)
    page = pdf_reader.pages[0]
    text_pdf = page.extract_text()
    print(text_pdf)
    assert number_of_page == 1
    assert pdf_size == 20597
    assert 'PDF Test File' in text_pdf

def test_check_csv(create_zip):
    with open('./hidden/resources/SampleCSVFile_11kb.csv') as csvfile:
        table = csv.reader(csvfile)
        for row, line in enumerate(table, 1):
            if row == 5:
                assert line[1] == 'Holmes HEPA Air Purifier'
    csv_size = os.path.getsize('./hidden/resources/SampleCSVFile_11kb.csv')
    assert csv_size == 10998

def test_check_xlsx(create_zip):
    workbook = load_workbook('./hidden/resources/file_example_XLSX_10.xlsx')
    sheet = workbook.active
    text_xlsx = sheet.cell(row=3, column=5).value
    xlsx_size = os.path.getsize('./hidden/resources/file_example_XLSX_10.xlsx')
    assert "Great Britain" in text_xlsx
    assert xlsx_size == 5425
